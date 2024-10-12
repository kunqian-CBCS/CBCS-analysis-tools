import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
from sklearn.metrics import auc
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
import win32com.client

# Load the data files
file_24h = r'C:\Users\kun.qian\Desktop\Projects\U2OS phospholipidoses assay\Elin dose and time points\Elin_U2OS_PL\U2OS_finalPreparedDR_24h_50cutoff.xlsx'
file_72h = r'C:\Users\kun.qian\Desktop\Projects\U2OS phospholipidoses assay\Elin dose and time points\Elin_U2OS_PL\U2OS_finalPreparedDR_72h_50cutoff.xlsx'
file_ic50_initial = r"C:\Users\kun.qian\Desktop\Projects\U2OS phospholipidoses assay\Elin dose and time points\Elin_U2OS_PL\U2OS_initial_guesses.xlsx"

data_24h = pd.read_excel(file_24h)
data_72h = pd.read_excel(file_72h)
ic50_initial = pd.read_excel(file_ic50_initial)

# Add time point information
data_24h['time'] = 24
data_72h['time'] = 72

# Standardize column names
data_72h.columns = data_24h.columns

# Combine data into a single DataFrame
combined_data = pd.concat([data_24h, data_72h])

# Rename columns in ic50_initial to match the format we need
ic50_initial = ic50_initial.rename(columns={
    'IC50_24': 'IC50_24h', 
    'Slope_24': 'Slope_24h',
    'IC50_72': 'IC50_72h', 
    'Slope_72': 'Slope_72h'
})

# Convert initial IC50 guesses from M to nM
ic50_initial['IC50_24h'] *= 1e9
ic50_initial['IC50_72h'] *= 1e9

# Merge initial IC50 and slope guesses with combined data
combined_data = combined_data.merge(ic50_initial, on='Batch_nr', how='left')

# Function to fit dose-response curve and calculate IC50 and AUC
def dose_response_curve(conc, inhib, time, batch, initial_ic50, initial_slope):
    if len(conc) < 4:
        print(f"Not enough data points for batch {batch} at {time}h.")
        return np.nan, np.nan
    
    def logistic_model(x, A, B, C, D):
        return A + (B - A) / (1.0 + (C / x)**D)
    
    # Use the initial IC50 and slope values as the starting points for curve fitting
    initial_guesses = [np.min(inhib), np.max(inhib), initial_ic50, initial_slope]
    
    try:
        popt, _ = curve_fit(logistic_model, conc, inhib, p0=initial_guesses, maxfev=100000)
        A, B, C, D = popt
        ic50 = C
        # Generate 500 evenly spaced values between min and max concentration for a smooth curve
        x_vals = np.linspace(min(conc), max(conc), 500)
        y_vals = logistic_model(x_vals, *popt)
        area = auc(x_vals, y_vals)
        
        # Plot the smooth curve
        plt.plot(x_vals, y_vals, label=f'{time}h (IC50={ic50:.2f}, AUC={area:.2f})')
        
        # Plot the data points
        plt.scatter(conc, inhib, color=plt.gca().lines[-1].get_color(), s=50, zorder=5)  # Same color as the curve, size 50
        
        return ic50, area
    except RuntimeError:
        print(f"Optimal parameters not found for batch {batch} at {time}h.")
        return np.nan, np.nan

# Create a new Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "combined_time_points"

# Write the header row
header = [
    'Batch_nr', 'IC50_24h', 'IC50_72h', 'AUC_24h', 'AUC_72h', 'GRAPH'
]
ws.append(header)

# Adjust column width for the 'GRAPH' column
graph_col_letter = get_column_letter(len(header))
ws.column_dimensions[graph_col_letter].width = 40  # Adjust as needed

# Save the initial workbook
output_dir = r'C:\Users\kun.qian\Desktop\Projects\U2OS phospholipidoses assay\Elin dose and time points\Elin_U2OS_PL'
figures_dir = os.path.join(output_dir, 'figures')
os.makedirs(figures_dir, exist_ok=True)
excel_path = os.path.join(output_dir, 'U2OS_combined_time_points_with_plots_50cutoff.xlsx')
wb.save(excel_path)

# Plotting function
def plot_ic50_curve(batch_data, batch, ws, row_idx):
    plt.figure(figsize=(6, 4))  # Adjust the figure size as needed
    ic50_values = []
    auc_values = []
    
    for time in [24, 72]:
        data_subset = batch_data[batch_data['time'] == time]
        initial_ic50 = data_subset[f'IC50_{time}h'].iloc[0]
        initial_slope = data_subset[f'Slope_{time}h'].iloc[0]
        ic50, area = dose_response_curve(data_subset['Conc_nM'], data_subset['inhibition'], time, batch, initial_ic50, initial_slope)
        ic50_values.append(ic50)
        auc_values.append(area)
    
    plt.xscale('log')
    plt.xlabel('Concentration')
    plt.ylabel('Response')
    plt.title(f'Dose-Response Curves for Batch {batch}')
    plt.legend()
    plt.grid(True)
    
    # Save the plot to a file in the figures folder
    plot_filename = os.path.join(figures_dir, f'{batch}_ic50_curve.png')
    plt.savefig(plot_filename, bbox_inches='tight')
    plt.close()
    
    # Append IC50 and AUC values to the worksheet
    ws.append([batch] + ic50_values + auc_values + [''])
    
    return plot_filename, row_idx

# Generate plots for each batch
plot_filenames = []
batches = combined_data['Batch_nr'].unique()
for batch in batches:
    batch_data = combined_data[combined_data['Batch_nr'] == batch]
    row_idx = ws.max_row + 1
    plot_filename, row_idx = plot_ic50_curve(batch_data, batch, ws, row_idx)
    plot_filenames.append((plot_filename, row_idx))

# Save the workbook
wb.save(excel_path)

# Insert images into the Excel workbook using win32com.client
Excel = win32com.client.Dispatch("Excel.Application")
Excel.Visible = False
wb_win32 = Excel.Workbooks.Open(excel_path)
ws_win32 = wb_win32.Worksheets("combined_time_points")

# Find the column named "GRAPH"
graph_col = None
for col in range(1, ws_win32.UsedRange.Columns.Count + 1):
    if ws_win32.Cells(1, col).Value == 'GRAPH':
        graph_col = col
        break

if graph_col is None:
    raise ValueError("The column 'GRAPH' was not found in the Excel sheet.")

for pic_path, row_idx in plot_filenames:
    # Verify that the image file exists
    if not os.path.exists(pic_path):
        print(f"File not found: {pic_path}")
        continue
    
    # Debug statements
    print(f"Inserting image: {pic_path}")
    print(f"Row index: {row_idx}")

    # Set row height
    ws_win32.Rows(row_idx).RowHeight = 120  # Adjust as needed
    
    left = ws_win32.Cells(row_idx, graph_col).Left
    top = ws_win32.Cells(row_idx, graph_col).Top
    col_width = ws_win32.Columns(graph_col).ColumnWidth * 7.5  # Approximate width in points
    img_width, img_height = plt.imread(pic_path).shape[1], plt.imread(pic_path).shape[0]
    aspect_ratio = img_width / img_height
    width = col_width
    height = width / aspect_ratio

    if height > ws_win32.Rows(row_idx).RowHeight:
        height = ws_win32.Rows(row_idx).RowHeight
        width = height * aspect_ratio
    
    print(f"Position: left={left}, top={top}, width={width}, height={height}")
    
    ws_win32.Shapes.AddPicture(pic_path, LinkToFile=False, SaveWithDocument=True, Left=left, Top=top, Width=width, Height=height)

wb_win32.Save()
wb_win32.Close()
Excel.Quit()
