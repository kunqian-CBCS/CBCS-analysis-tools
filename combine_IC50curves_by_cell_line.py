import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import win32com.client

# Define the base directory and file paths for the four Excel files
base_dir = r'C:\Users\kun.qian\Desktop\Projects\Nordic Oncology Library\screening test\FiMMs comparison\python combined dose curves\four cell lines'
file_paths = {
    'HL60': os.path.join(base_dir, 'HL60_DSRT_analysis_table_Rpipeline_IC50.xlsx'),
    'Kuramochi': os.path.join(base_dir, 'Kuramochi_DSRT_analysis_table_Rpipeline_IC50.xlsx'),
    'MOLM13': os.path.join(base_dir, 'MOLM13_DSRT_analysis_table_Rpipeline_IC50.xlsx'),
    'Ovcar8': os.path.join(base_dir, 'Ovcar8_DSRT_analysis_table_Rpipeline_IC50.xlsx')
}

# Load data from each file into a dictionary of DataFrames
data_frames = {}
for cell_line, path in file_paths.items():
    df = pd.read_excel(path)
    df['Cell_Line'] = cell_line  # Add a column to identify the cell line
    data_frames[cell_line] = df

# Combine the data into a single DataFrame
data = pd.concat(data_frames.values(), ignore_index=True)

# Extract relevant columns, assuming the column 'Cell_Line' identifies different cell lines
data = data[['ID', 'DRUG_NAME', 'Cell_Line', 'D1', 'D2', 'D3', 'D4', 'D5', 'IC50', 'DSS', 'SLOPE', 'MAX', 'MIN', 'Max.Conc.tested']]

# Remove rows with any inf or NaN values
data = data.replace([np.inf, -np.inf], np.nan).dropna()

# Create a function to fit the Hill equation
def hill_equation(concentration, ic50, hill_slope, min_resp, max_resp):
    concentration = np.clip(concentration, 1e-10, np.inf)  # Avoid extremely small values
    return min_resp + (max_resp - min_resp) / (1 + np.power(concentration / ic50, hill_slope))

# Create output directory for plots
output_dir = os.path.join(base_dir, 'ic50_plots')
os.makedirs(output_dir, exist_ok=True)

# Create a new Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "IC50 Summary"

# Write the header row
header = [
    'Drug Name', 'HL60 IC50', 'Kuramochi IC50', 'MOLM13 IC50', 'Ovcar8 IC50',
    'HL60 DSS', 'Kuramochi DSS', 'MOLM13 DSS', 'Ovcar8 DSS',
    'HL60 IC50_calc', 'Kuramochi IC50_calc', 'MOLM13 IC50_calc', 'Ovcar8 IC50_calc', 'GRAPH'
]
ws.append(header)

# Adjust column width for the 'GRAPH' column
graph_col_letter = get_column_letter(len(header))
ws.column_dimensions[graph_col_letter].width = 40  # Adjust as needed

# Save the initial workbook
initial_excel_path = os.path.join(base_dir, 'IC50_summary_with_plots.xlsx')
wb.save(initial_excel_path)

# Plotting function
def plot_ic50_curve(drug_data, drug_name):
    plt.figure(figsize=(6, 4))  # Adjust the figure size as needed
    ic50_values = []
    dss_values = []
    ic50_calc_values = []
    
    for cell_line in drug_data['Cell_Line'].unique():
        cell_line_data = drug_data[drug_data['Cell_Line'] == cell_line]
        
        for index, row in cell_line_data.iterrows():
            # Calculate concentrations for D1, D2, D3, D4, and D5 based on max conc only
            max_conc = row['Max.Conc.tested']
            concentrations_tested = np.array([max_conc / 10000, max_conc / 1000, max_conc / 100, max_conc / 10, max_conc])
            data_points = row[['D1', 'D2', 'D3', 'D4', 'D5']].values
            
            # Initial guesses for curve fitting
            initial_guesses = [row['IC50'], row['SLOPE'], row['MIN'], row['MAX']]
            
            # Fit the Hill equation to the data points
            try:
                popt, _ = curve_fit(hill_equation, concentrations_tested, data_points, p0=initial_guesses, bounds=(0, [np.inf, np.inf, np.inf, np.inf]), maxfev=2000)
                ic50, hill_slope, min_resp, max_resp = popt
                
                # Generate smooth curve for plotting
                concentrations_smooth = np.logspace(np.log10(concentrations_tested[0]), np.log10(concentrations_tested[-1]), 100)
                response_smooth = hill_equation(concentrations_smooth, ic50, hill_slope, min_resp, max_resp)
                
                color = next(plt.gca()._get_lines.prop_cycler)['color']  # Get the current color
                
                plt.plot(concentrations_smooth, response_smooth, label=f"{cell_line} - {row['ID']}", color=color)
                # Plotting the actual data points as dots
                plt.scatter(concentrations_tested, data_points, marker='o', color=color)
                ic50_values.append(row['IC50'])
                dss_values.append(row['DSS'])
                ic50_calc_values.append(ic50)
            except RuntimeError:
                print(f"Optimal parameters not found for {drug_name} in {cell_line} - {row['ID']}")
                ic50_values.append(np.nan)
                dss_values.append(row['DSS'])
                ic50_calc_values.append(np.nan)
            except ValueError:
                print(f"Invalid values encountered for {drug_name} in {cell_line} - {row['ID']}")
                ic50_values.append(np.nan)
                dss_values.append(row['DSS'])
                ic50_calc_values.append(np.nan)
    
    plt.xscale('log')
    plt.xlabel('Concentration')
    plt.ylabel('Response')
    plt.title(f'IC50 Curves for {drug_name}')
    plt.legend()
    plt.grid(True)
    
    # Save the plot to a file
    plot_filename = os.path.join(output_dir, f'{drug_name}_ic50_curve.png')
    plt.savefig(plot_filename, bbox_inches='tight')
    plt.close()
    
    return plot_filename, ic50_values, dss_values, ic50_calc_values

# Generate plots for each drug
plot_filenames = []
unique_drugs = data['DRUG_NAME'].unique()
for drug in unique_drugs:
    drug_data = data[data['DRUG_NAME'] == drug]
    plot_filename, ic50_values, dss_values, ic50_calc_values = plot_ic50_curve(drug_data, drug)
    
    # Append IC50, DSS values, and calculated IC50 to the worksheet
    row_idx = ws.max_row + 1
    ws.append([drug] + ic50_values + dss_values + ic50_calc_values + [''])  # Append calculated IC50 values
    plot_filenames.append((plot_filename, row_idx))

# Save the workbook
wb.save(initial_excel_path)

# Insert images into the Excel workbook using win32com.client
Excel = win32com.client.Dispatch("Excel.Application")
Excel.Visible = False
wb_win32 = Excel.Workbooks.Open(initial_excel_path)
ws_win32 = wb_win32.Worksheets("IC50 Summary")

# Find the column named "GRAPH"
graph_col = None
for col in range(1, ws_win32.UsedRange.Columns.Count + 1):
    if ws_win32.Cells(1, col).Value == 'GRAPH':
        graph_col = col
        break

if graph_col is None:
    raise ValueError("The column 'GRAPH' was not found in the Excel sheet.")

for pic_path, row_idx in plot_filenames:
    # Verify that the image file exists
    if not os.path.exists(pic_path):
        print(f"File not found: {pic_path}")
        continue
    
    # Debug statements
    print(f"Inserting image: {pic_path}")
    print(f"Row index: {row_idx}")

    # Set row height
    ws_win32.Rows(row_idx).RowHeight = 120  # Adjust as needed
    
    left = ws_win32.Cells(row_idx, graph_col).Left
    top = ws_win32.Cells(row_idx, graph_col).Top
    col_width = ws_win32.Columns(graph_col).ColumnWidth * 7.5  # Approximate width in points
    img_width, img_height = plt.imread(pic_path).shape[1], plt.imread(pic_path).shape[0]
    aspect_ratio = img_width / img_height
    width = col_width
    height = width / aspect_ratio

    if height > ws_win32.Rows(row_idx).RowHeight:
        height = ws_win32.Rows(row_idx).RowHeight
        width = height * aspect_ratio
    
    print(f"Position: left={left}, top={top}, width={width}, height={height}")
    
    ws_win32.Shapes.AddPicture(pic_path, LinkToFile=False, SaveWithDocument=True, Left=left, Top=top, Width=width, Height=height)

wb_win32.Save()
wb_win32.Close()
Excel.Quit()
